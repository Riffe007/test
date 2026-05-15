#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""TFLite vs PyTorch (qfgaohao) MobileNetV2-SSD comparison on COCO val2017.

Runs both models on the first N COCO val2017 images and writes a side-by-side
xlsx comparison.

Models
------
* tf2_ssd_mobilenet_v2_coco17_ptq.tflite -- COCO 80-class, decode baked in via
  TFLite_Detection_PostProcess.
* qfgaohao MobileNetV2-SSD-Lite -- VOC 20-class, runs through qfgaohao's
  reference ``Predictor`` (is_test=True path) so this script is independent
  of the toolkit's Python-side decode.

Fair comparison
---------------
The two models target different label spaces. To keep the comparison
meaningful both are scored on the same 20 VOC-equivalent COCO category IDs
via :data:`VOC_INDEX_TO_COCO_ID`. The TFLite model also gets a full-80-class
mAP for context.

Output
------
An xlsx with four sheets:
* ``Summary``           -- overall metrics + latency stats (formulas)
* ``Per-Class``         -- per-class AP@0.5 for the 20 VOC-equivalent classes
* ``Per-Image``         -- one row per image: counts, max score, latency
* ``Config``            -- run parameters and model paths

Usage
-----
::

    python compare_tflite_pytorch_coco.py \\
        --tflite-model     <path/to/.tflite> \\
        --pytorch-model    <path/to/.pth> \\
        --qfgaohao-repo    <path/to/pytorch-ssd> \\
        --coco-images      <path/to/val2017> \\
        --coco-annotations <path/to/instances_val2017.json> \\
        --output-xlsx      <path/to/out.xlsx> \\
        --num-images       1000

All paths default to the conventional locations under the MetaExecuTorch
project tree; see :func:`parse_args`.
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

import numpy as np
import torch
from PIL import Image, ImageDraw, ImageFont

# ---------------------------------------------------------------------------
# Third-party imports with graceful fallback messaging.
# ---------------------------------------------------------------------------

try:
    from tflite_runtime.interpreter import Interpreter as TFLiteInterpreter
except ImportError:
    try:
        import tensorflow as _tf  # type: ignore
        TFLiteInterpreter = _tf.lite.Interpreter  # type: ignore[attr-defined]
    except ImportError as _exc:  # pragma: no cover - import-time guard
        raise ImportError(
            "Need tflite-runtime or tensorflow for TFLite inference. "
            "Install one of:\n"
            "    pip install tflite-runtime\n"
            "    pip install tensorflow"
        ) from _exc

from pycocotools.coco import COCO
from pycocotools.cocoeval import COCOeval

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Class-space mappings.
# ---------------------------------------------------------------------------

# VOC class index (1-20, qfgaohao convention) -> COCO category_id.
# Used to project qfgaohao predictions into the COCO category space so both
# models can be scored against the same COCO ground truth.
VOC_INDEX_TO_COCO_ID: dict[int, int] = {
    1: 5,    # aeroplane     -> airplane
    2: 2,    # bicycle
    3: 16,   # bird
    4: 9,    # boat
    5: 44,   # bottle
    6: 6,    # bus
    7: 3,    # car
    8: 17,   # cat
    9: 62,   # chair
    10: 21,  # cow
    11: 67,  # diningtable   -> dining table
    12: 18,  # dog
    13: 19,  # horse
    14: 4,   # motorbike     -> motorcycle
    15: 1,   # person
    16: 64,  # pottedplant   -> potted plant
    17: 20,  # sheep
    18: 63,  # sofa          -> couch
    19: 7,   # train
    20: 72,  # tvmonitor     -> tv
}

VOC_INDEX_NAMES: dict[int, str] = {
    1:  "aeroplane",
    2:  "bicycle",
    3:  "bird",
    4:  "boat",
    5:  "bottle",
    6:  "bus",
    7:  "car",
    8:  "cat",
    9:  "chair",
    10: "cow",
    11: "diningtable",
    12: "dog",
    13: "horse",
    14: "motorbike",
    15: "person",
    16: "pottedplant",
    17: "sheep",
    18: "sofa",
    19: "train",
    20: "tvmonitor",
}

# COCO 80-class TFLite output index (0-79) -> COCO category_id.
# Standard mapping for tf2_ssd_mobilenet_v2_coco17_ptq.tflite and the rest of
# the Coral SSD MobileNet V2 family. Index N is the Nth entry in the standard
# 80-class COCO label list (which skips the 10 unused ids in 1-90).
TFLITE_COCO80_INDEX_TO_COCO_ID: tuple[int, ...] = (
    1,  2,  3,  4,  5,  6,  7,  8,  9, 10,
    11, 13, 14, 15, 16, 17, 18, 19, 20, 21,
    22, 23, 24, 25, 27, 28, 31, 32, 33, 34,
    35, 36, 37, 38, 39, 40, 41, 42, 43, 44,
    46, 47, 48, 49, 50, 51, 52, 53, 54, 55,
    56, 57, 58, 59, 60, 61, 62, 63, 64, 65,
    67, 70, 72, 73, 74, 75, 76, 77, 78, 79,
    80, 81, 82, 84, 85, 86, 87, 88, 89, 90,
)

# Subset of COCO category_ids that map to a VOC class (sorted for stable
# COCOeval ordering).
VOC_EQUIVALENT_COCO_IDS: tuple[int, ...] = tuple(
    sorted(VOC_INDEX_TO_COCO_ID.values())
)


# ---------------------------------------------------------------------------
# Logging.
# ---------------------------------------------------------------------------

def _setup_logger() -> logging.Logger:
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )
    return logging.getLogger("compare_tflite_pytorch")


logger = _setup_logger()


# ---------------------------------------------------------------------------
# TFLite evaluator.
# ---------------------------------------------------------------------------

class TFLiteSSDEvaluator:
    """Run a TFLite SSD MobileNet V2 model and emit COCO-format detections.

    The model bakes ``TFLite_Detection_PostProcess`` into its graph, so the
    interpreter returns decoded boxes/classes/scores/count directly -- no
    Python-side prior decode or NMS is required.
    """

    # Standard output order for TF SSD models (sorted by output index).
    _EXPECTED_NUM_OUTPUTS = 4

    def __init__(self, model_path: Path) -> None:
        self.model_path = Path(model_path)
        self._interpreter = TFLiteInterpreter(model_path=str(self.model_path))
        self._interpreter.allocate_tensors()
        self._input_details = self._interpreter.get_input_details()
        self._output_details = sorted(
            self._interpreter.get_output_details(),
            key=lambda d: d["index"],
        )

        if len(self._output_details) != self._EXPECTED_NUM_OUTPUTS:
            raise RuntimeError(
                f"Expected {self._EXPECTED_NUM_OUTPUTS} output tensors "
                f"(boxes, classes, scores, num_detections); got "
                f"{len(self._output_details)}"
            )

        input_shape = self._input_details[0]["shape"]
        if len(input_shape) != 4 or input_shape[3] != 3:
            raise ValueError(
                f"Expected NHWC 3-channel input; got shape {input_shape}"
            )
        self._input_height = int(input_shape[1])
        self._input_width = int(input_shape[2])
        self._input_dtype = self._input_details[0]["dtype"]

        logger.info(
            "TFLite model loaded: %s | input: %s %s",
            self.model_path.name,
            tuple(input_shape),
            self._input_dtype.__name__,
        )

    # -- preprocessing --------------------------------------------------

    def _preprocess(self, pil_image: Image.Image) -> np.ndarray:
        """Resize, cast to model dtype, add batch dim. NHWC layout."""
        img = pil_image.convert("RGB").resize(
            (self._input_width, self._input_height), Image.BILINEAR
        )
        arr = np.asarray(img)
        if self._input_dtype == np.uint8:
            return arr[np.newaxis, ...]
        # Float input: TF SSD float graphs expect [-1, 1] (the standard
        # MobileNet preprocessing). The PTQ uint8 graph is what we expect
        # here, so this branch is defensive only.
        return ((arr.astype(np.float32) / 127.5) - 1.0)[np.newaxis, ...]

    # -- inference ------------------------------------------------------

    def _invoke(
        self, pil_image: Image.Image
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, int]:
        """Run one inference. Returns (boxes, classes, scores, count).

        ``boxes`` are in [ymin, xmin, ymax, xmax] normalized to [0, 1].
        ``classes`` are 0-indexed into :data:`TFLITE_COCO80_INDEX_TO_COCO_ID`.
        """
        input_tensor = self._preprocess(pil_image)
        self._interpreter.set_tensor(
            self._input_details[0]["index"], input_tensor
        )
        self._interpreter.invoke()

        tensors = [
            self._interpreter.get_tensor(d["index"])
            for d in self._output_details
        ]
        boxes_t, classes_t, scores_t, count_t = tensors

        if boxes_t.ndim != 3 or boxes_t.shape[-1] != 4:
            raise RuntimeError(
                f"Unexpected boxes shape: {boxes_t.shape}; this likely means "
                f"the output tensors are in a non-standard order"
            )

        count = int(np.asarray(count_t).flat[0])
        boxes = boxes_t[0][:count]
        classes = classes_t[0][:count].astype(int)
        scores = scores_t[0][:count]
        return boxes, classes, scores, count

    # -- public API -----------------------------------------------------

    def detections_for_image(
        self,
        pil_image: Image.Image,
        image_id: int,
        original_width: int,
        original_height: int,
        score_threshold: float,
    ) -> tuple[list[dict[str, Any]], float]:
        """Run inference and return ``(coco_detection_dicts, latency_ms)``."""
        t0 = time.perf_counter()
        boxes, classes, scores, count = self._invoke(pil_image)
        latency_ms = (time.perf_counter() - t0) * 1000.0

        detections: list[dict[str, Any]] = []
        max_idx = len(TFLITE_COCO80_INDEX_TO_COCO_ID)
        for i in range(count):
            score = float(scores[i])
            if score < score_threshold:
                continue
            class_idx = int(classes[i])
            if not 0 <= class_idx < max_idx:
                continue
            category_id = TFLITE_COCO80_INDEX_TO_COCO_ID[class_idx]

            ymin, xmin, ymax, xmax = (float(v) for v in boxes[i])
            x1 = xmin * original_width
            y1 = ymin * original_height
            x2 = xmax * original_width
            y2 = ymax * original_height
            w = max(x2 - x1, 0.0)
            h = max(y2 - y1, 0.0)

            detections.append({
                "image_id":    int(image_id),
                "category_id": int(category_id),
                "bbox":        [x1, y1, w, h],
                "score":       score,
            })

        return detections, latency_ms


# ---------------------------------------------------------------------------
# PyTorch evaluator (qfgaohao Predictor).
# ---------------------------------------------------------------------------

class PyTorchSSDEvaluator:
    """Run qfgaohao MobileNetV2-SSD-Lite using its reference ``Predictor``.

    We use ``is_test=True`` and the qfgaohao predictor end-to-end so this
    script is independent of any in-toolkit decode logic. Predictor output
    classes (1-20 VOC indices) are remapped to COCO category_ids via
    :data:`VOC_INDEX_TO_COCO_ID`.
    """

    def __init__(
        self,
        model_path: Path,
        qfgaohao_repo_path: Path,
        num_classes: int = 21,
        device: str = "cpu",
        candidate_size: int = 200,
    ) -> None:
        self.model_path = Path(model_path)
        self.repo_path = Path(qfgaohao_repo_path).resolve()
        self.device = device

        if str(self.repo_path) not in sys.path:
            sys.path.insert(0, str(self.repo_path))

        from vision.ssd.mobilenetv2_ssd_lite import (  # type: ignore
            create_mobilenetv2_ssd_lite,
            create_mobilenetv2_ssd_lite_predictor,
        )

        self._net = create_mobilenetv2_ssd_lite(num_classes, is_test=True)
        self._load_weights_into(self._net, self.model_path)
        self._net.eval()

        self._predictor = create_mobilenetv2_ssd_lite_predictor(
            self._net,
            candidate_size=candidate_size,
            device=torch.device(device),
        )
        logger.info("PyTorch model loaded: %s", self.model_path.name)

    @staticmethod
    def _load_checkpoint(path: Path) -> Any:
        """Load a torch checkpoint with weights_only fallback for old torch."""
        try:
            return torch.load(str(path), map_location="cpu", weights_only=False)
        except TypeError:
            # Older torch without weights_only kwarg.
            return torch.load(str(path), map_location="cpu")

    @classmethod
    def _load_weights_into(
        cls, net: torch.nn.Module, model_path: Path
    ) -> None:
        """Handle the two pickle shapes qfgaohao checkpoints come in.

        * Full ``nn.Module``: copy its ``state_dict()`` into the freshly
          built net.
        * Dict with ``state_dict`` key (or a bare state_dict): load directly.
        """
        ckpt = cls._load_checkpoint(model_path)
        if isinstance(ckpt, torch.nn.Module):
            net.load_state_dict(ckpt.state_dict())
        elif isinstance(ckpt, dict) and "state_dict" in ckpt:
            net.load_state_dict(ckpt["state_dict"])
        elif isinstance(ckpt, dict):
            net.load_state_dict(ckpt)
        else:
            raise TypeError(
                f"Unsupported checkpoint type for {model_path}: {type(ckpt)}"
            )

    def detections_for_image(
        self,
        pil_image: Image.Image,
        image_id: int,
        original_width: int,  # noqa: ARG002 -- predictor handles its own scale
        original_height: int,  # noqa: ARG002
        score_threshold: float,
    ) -> tuple[list[dict[str, Any]], float]:
        """Run ``Predictor.predict`` and return ``(coco_dicts, latency_ms)``.

        qfgaohao's predictor returns boxes already scaled to the input
        image's pixel coordinate space, so no further denormalization is
        required.
        """
        # Predictor expects HWC numpy uint8 RGB.
        np_image = np.asarray(pil_image.convert("RGB"))

        t0 = time.perf_counter()
        boxes, labels, probs = self._predictor.predict(
            np_image, top_k=-1, prob_threshold=score_threshold
        )
        latency_ms = (time.perf_counter() - t0) * 1000.0

        detections: list[dict[str, Any]] = []
        if boxes is None or (hasattr(boxes, "numel") and boxes.numel() == 0):
            return detections, latency_ms

        boxes_np = _to_numpy(boxes)
        labels_np = _to_numpy(labels).astype(int)
        probs_np = _to_numpy(probs)

        for i in range(len(boxes_np)):
            voc_idx = int(labels_np[i])
            category_id = VOC_INDEX_TO_COCO_ID.get(voc_idx)
            if category_id is None:
                continue
            x1, y1, x2, y2 = (float(v) for v in boxes_np[i])
            w = max(x2 - x1, 0.0)
            h = max(y2 - y1, 0.0)
            detections.append({
                "image_id":    int(image_id),
                "category_id": int(category_id),
                "bbox":        [x1, y1, w, h],
                "score":       float(probs_np[i]),
            })

        return detections, latency_ms


def _to_numpy(t: Any) -> np.ndarray:
    """Coerce torch tensor or array-like to numpy."""
    if torch.is_tensor(t):
        return t.detach().cpu().numpy()
    return np.asarray(t)


# ---------------------------------------------------------------------------
# COCO evaluation.
# ---------------------------------------------------------------------------

_METRIC_KEYS: tuple[str, ...] = (
    "mAP_0.5_0.95",
    "mAP_0.5",
    "mAP_0.75",
    "mAP_small",
    "mAP_medium",
    "mAP_large",
    "AR_max1",
    "AR_max10",
    "AR_max100",
)


def _empty_metrics() -> dict[str, float]:
    return {k: 0.0 for k in _METRIC_KEYS}


def run_coco_eval(
    coco_gt: COCO,
    predictions: list[dict[str, Any]],
    category_ids: Optional[Iterable[int]] = None,
    image_ids: Optional[Iterable[int]] = None,
) -> tuple[dict[str, float], dict[int, float]]:
    """One COCOeval pass returning ``(summary_metrics, per_class_ap_at_0.5)``.

    The per-class dict only covers categories in ``category_ids`` (or all
    GT categories if ``None``).
    """
    if category_ids is not None:
        cat_list: list[int] = list(category_ids)
    else:
        cat_list = sorted(coco_gt.getCatIds())

    summary = _empty_metrics()
    per_class: dict[int, float] = {cid: 0.0 for cid in cat_list}

    if not predictions:
        return summary, per_class

    try:
        coco_dt = coco_gt.loadRes(predictions)
    except Exception as exc:
        logger.error("COCO loadRes failed: %s", exc)
        return summary, per_class

    coco_eval = COCOeval(coco_gt, coco_dt, "bbox")
    coco_eval.params.catIds = cat_list
    if image_ids is not None:
        coco_eval.params.imgIds = list(image_ids)

    coco_eval.evaluate()
    coco_eval.accumulate()
    coco_eval.summarize()

    stats = coco_eval.stats
    summary = {
        "mAP_0.5_0.95": float(stats[0]),
        "mAP_0.5":      float(stats[1]),
        "mAP_0.75":     float(stats[2]),
        "mAP_small":    float(stats[3]),
        "mAP_medium":   float(stats[4]),
        "mAP_large":    float(stats[5]),
        "AR_max1":      float(stats[6]),
        "AR_max10":     float(stats[7]),
        "AR_max100":    float(stats[8]),
    }

    # Per-class AP@0.5 from coco_eval.eval['precision'].
    # Shape: [T iouThrs, R recallThrs, K cats, A areas, M maxDets].
    precision = getattr(coco_eval, "eval", {}).get("precision")
    if precision is not None and precision.size > 0:
        for k, cid in enumerate(cat_list):
            slc = precision[0, :, k, 0, -1]  # IoU=0.5, all R, this cat
            valid = slc[slc > -1]
            per_class[cid] = float(valid.mean()) if valid.size else 0.0

    return summary, per_class


# ---------------------------------------------------------------------------
# Debug image rendering.
# ---------------------------------------------------------------------------

# Colors per source. Outline + label background share the same hue.
_DEBUG_COLOR_GT       = (46, 184, 46)    # green
_DEBUG_COLOR_TFLITE   = (52, 134, 235)   # blue
_DEBUG_COLOR_PYTORCH  = (228, 52, 52)    # red

# Common TTF locations on Ubuntu / WSL. First hit wins; fall back to PIL's
# bitmap default which is tiny but always available.
_CANDIDATE_FONT_PATHS: tuple[str, ...] = (
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/mnt/c/Windows/Fonts/arialbd.ttf",
    "/mnt/c/Windows/Fonts/arial.ttf",
)


def _load_label_font(size: int = 14) -> ImageFont.ImageFont:
    """Try a handful of common TTF paths, fall back to PIL's default."""
    for path in _CANDIDATE_FONT_PATHS:
        try:
            return ImageFont.truetype(path, size=size)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


# Cache the font once - cheap to share across calls.
_LABEL_FONT_CACHE: Optional[ImageFont.ImageFont] = None
_TITLE_FONT_CACHE: Optional[ImageFont.ImageFont] = None


def _label_font() -> ImageFont.ImageFont:
    global _LABEL_FONT_CACHE
    if _LABEL_FONT_CACHE is None:
        _LABEL_FONT_CACHE = _load_label_font(size=14)
    return _LABEL_FONT_CACHE


def _title_font() -> ImageFont.ImageFont:
    global _TITLE_FONT_CACHE
    if _TITLE_FONT_CACHE is None:
        _TITLE_FONT_CACHE = _load_label_font(size=20)
    return _TITLE_FONT_CACHE


def _text_size(
    draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont
) -> tuple[int, int]:
    """Cross-version Pillow text measurement.

    Pillow >= 8 deprecated ``textsize`` in favor of ``textbbox``; this helper
    keeps the script compatible with both.
    """
    if hasattr(draw, "textbbox"):
        left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
        return right - left, bottom - top
    return draw.textsize(text, font=font)  # type: ignore[attr-defined]


def _draw_box_with_label(
    draw: ImageDraw.ImageDraw,
    xywh: Sequence[float],
    label: str,
    color: tuple[int, int, int],
    image_size: tuple[int, int],
    *,
    outline_width: int = 3,
) -> None:
    """Draw a single (box, label) pair clipped to the image extent."""
    img_w, img_h = image_size
    x, y, w, h = xywh
    x1 = max(0.0, float(x))
    y1 = max(0.0, float(y))
    x2 = min(float(img_w), float(x) + float(w))
    y2 = min(float(img_h), float(y) + float(h))
    if x2 <= x1 or y2 <= y1:
        return

    draw.rectangle([x1, y1, x2, y2], outline=color, width=outline_width)

    font = _label_font()
    text_w, text_h = _text_size(draw, label, font)

    # Place the label inside the box near the top-left, but flip below the
    # box edge if it would otherwise extend above the image.
    label_y = y1 - text_h - 2
    if label_y < 0:
        label_y = y1 + 2

    # Filled background strip so text is legible regardless of image content.
    draw.rectangle(
        [x1, label_y, x1 + text_w + 6, label_y + text_h + 2],
        fill=color,
    )
    draw.text((x1 + 3, label_y), label, fill=(255, 255, 255), font=font)


def _render_panel(
    base_image: Image.Image,
    boxes_xywh: Sequence[Sequence[float]],
    labels: Sequence[str],
    color: tuple[int, int, int],
    title: str,
) -> Image.Image:
    """Render a single titled panel with overlaid boxes."""
    panel = base_image.copy().convert("RGB")
    draw = ImageDraw.Draw(panel)

    for box, label in zip(boxes_xywh, labels):
        _draw_box_with_label(draw, box, label, color, panel.size)

    # Title strip at the top.
    title_font = _title_font()
    title_w, title_h = _text_size(draw, title, title_font)
    strip_h = title_h + 8
    draw.rectangle([0, 0, panel.width, strip_h], fill=color)
    draw.text(
        ((panel.width - title_w) // 2, 4),
        title,
        fill=(255, 255, 255),
        font=title_font,
    )
    return panel


def _stitch_horizontal(
    panels: Sequence[Image.Image], gap: int = 8, bg: tuple[int, int, int] = (32, 32, 32)
) -> Image.Image:
    """Concatenate panels left-to-right with a small gap."""
    if not panels:
        raise ValueError("Need at least one panel to stitch.")
    height = max(p.height for p in panels)
    width = sum(p.width for p in panels) + gap * (len(panels) - 1)
    canvas = Image.new("RGB", (width, height), bg)
    x = 0
    for p in panels:
        canvas.paste(p, (x, 0))
        x += p.width + gap
    return canvas


def render_debug_image(
    pil_image: Image.Image,
    *,
    gt_anns: Sequence[dict[str, Any]],
    tflite_dets: Sequence[dict[str, Any]],
    pytorch_dets: Sequence[dict[str, Any]],
    category_name_by_id: dict[int, str],
    save_path: Path,
    display_score_threshold: float,
) -> None:
    """Write a 3-panel side-by-side image (GT | TFLite | qfgaohao) to disk.

    Each panel shows the same source image with boxes from one source
    overlaid in a distinctive color.
    """
    image_rgb = pil_image.convert("RGB")

    # GT panel: every annotation, no score filter.
    gt_boxes = [ann["bbox"] for ann in gt_anns]
    gt_labels = [
        category_name_by_id.get(int(ann["category_id"]), str(ann["category_id"]))
        for ann in gt_anns
    ]
    gt_panel = _render_panel(
        image_rgb, gt_boxes, gt_labels,
        color=_DEBUG_COLOR_GT,
        title=f"GROUND TRUTH ({len(gt_anns)})",
    )

    # Detection panels: filter by display threshold so the visualization
    # isn't choked by low-score noise.
    def _filter(dets: Sequence[dict[str, Any]]) -> tuple[list[list[float]], list[str]]:
        kept_boxes: list[list[float]] = []
        kept_labels: list[str] = []
        for d in dets:
            if float(d.get("score", 0.0)) < display_score_threshold:
                continue
            kept_boxes.append(list(d["bbox"]))
            name = category_name_by_id.get(
                int(d["category_id"]), str(d["category_id"])
            )
            kept_labels.append(f"{name}:{float(d['score']):.2f}")
        return kept_boxes, kept_labels

    tf_boxes, tf_labels = _filter(tflite_dets)
    pt_boxes, pt_labels = _filter(pytorch_dets)

    tflite_panel = _render_panel(
        image_rgb, tf_boxes, tf_labels,
        color=_DEBUG_COLOR_TFLITE,
        title=f"TFLite ({len(tf_boxes)}/{len(tflite_dets)})",
    )
    pytorch_panel = _render_panel(
        image_rgb, pt_boxes, pt_labels,
        color=_DEBUG_COLOR_PYTORCH,
        title=f"PyTorch / qfgaohao ({len(pt_boxes)}/{len(pytorch_dets)})",
    )

    combined = _stitch_horizontal([gt_panel, tflite_panel, pytorch_panel])
    save_path.parent.mkdir(parents=True, exist_ok=True)
    combined.save(save_path, format="PNG", optimize=True)


# ---------------------------------------------------------------------------
# XLSX writer.
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="305496")
_LABEL_FONT = Font(name="Arial", bold=True)
_TITLE_FONT = Font(name="Arial", bold=True, size=14)
_CENTER = Alignment(horizontal="center", vertical="center")


def _style_header_row(sheet: Worksheet, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _autosize_columns(
    sheet: Worksheet, min_width: int = 12, max_width: int = 60
) -> None:
    for col_cells in sheet.columns:
        # Skip empty columns (can happen on merged cells in oldest openpyxl).
        first = next((c for c in col_cells if c.value is not None), None)
        if first is None:
            continue
        col_letter = get_column_letter(first.column)
        longest = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        sheet.column_dimensions[col_letter].width = max(
            min_width, min(max_width, longest + 2)
        )


def write_xlsx(  # noqa: PLR0913 -- report writer is intrinsically wide
    output_xlsx: Path,
    *,
    num_images_evaluated: int,
    tflite_full_metrics: dict[str, float],
    tflite_voc_metrics: dict[str, float],
    pytorch_voc_metrics: dict[str, float],
    tflite_per_class: dict[int, float],
    pytorch_per_class: dict[int, float],
    per_image_rows: list[dict[str, Any]],
    tflite_model_path: Path,
    pytorch_model_path: Path,
    coco_gt_json: Path,
    coco_images_dir: Path,
    score_threshold: float,
) -> None:
    """Write the four-sheet comparison workbook."""
    wb = Workbook()

    _write_summary_sheet(
        wb.active,
        num_images_evaluated=num_images_evaluated,
        tflite_full=tflite_full_metrics,
        tflite_voc=tflite_voc_metrics,
        pytorch_voc=pytorch_voc_metrics,
        per_image_count=len(per_image_rows),
        score_threshold=score_threshold,
    )
    _write_per_class_sheet(
        wb.create_sheet("Per-Class"),
        tflite_per_class=tflite_per_class,
        pytorch_per_class=pytorch_per_class,
    )
    _write_per_image_sheet(wb.create_sheet("Per-Image"), per_image_rows)
    _write_config_sheet(
        wb.create_sheet("Config"),
        tflite_model_path=tflite_model_path,
        pytorch_model_path=pytorch_model_path,
        coco_gt_json=coco_gt_json,
        coco_images_dir=coco_images_dir,
        score_threshold=score_threshold,
        num_images_evaluated=num_images_evaluated,
    )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    logger.info("Wrote %s", output_xlsx)


def _write_summary_sheet(
    sheet: Worksheet,
    *,
    num_images_evaluated: int,
    tflite_full: dict[str, float],
    tflite_voc: dict[str, float],
    pytorch_voc: dict[str, float],
    per_image_count: int,
    score_threshold: float,
) -> None:
    sheet.title = "Summary"

    sheet["A1"] = "TFLite vs PyTorch (qfgaohao) — MobileNetV2-SSD on COCO val2017"
    sheet["A1"].font = _TITLE_FONT
    sheet.merge_cells("A1:D1")
    sheet["A1"].alignment = _CENTER

    sheet["A3"] = "Images evaluated:"
    sheet["A3"].font = _LABEL_FONT
    sheet["B3"] = num_images_evaluated
    sheet["A4"] = "Score threshold:"
    sheet["A4"].font = _LABEL_FONT
    sheet["B4"] = score_threshold

    header_row = 6
    headers = (
        "Metric",
        "TFLite (full COCO 80)",
        "TFLite (VOC subset)",
        "PyTorch (VOC subset)",
    )
    for col, h in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col, value=h)
    _style_header_row(sheet, header_row, len(headers))

    metric_rows: Sequence[tuple[str, str]] = (
        ("mAP @ IoU=0.5",         "mAP_0.5"),
        ("mAP @ IoU=0.75",        "mAP_0.75"),
        ("mAP @ IoU=[0.5:0.95]",  "mAP_0.5_0.95"),
        ("mAP (small objects)",   "mAP_small"),
        ("mAP (medium objects)",  "mAP_medium"),
        ("mAP (large objects)",   "mAP_large"),
        ("AR @ maxDets=1",        "AR_max1"),
        ("AR @ maxDets=10",       "AR_max10"),
        ("AR @ maxDets=100",      "AR_max100"),
    )
    row = header_row + 1
    for label, key in metric_rows:
        sheet.cell(row=row, column=1, value=label).font = _LABEL_FONT
        sheet.cell(row=row, column=2, value=tflite_full[key]).number_format = "0.0000"
        sheet.cell(row=row, column=3, value=tflite_voc[key]).number_format = "0.0000"
        sheet.cell(row=row, column=4, value=pytorch_voc[key]).number_format = "0.0000"
        row += 1

    # Latency block. Uses Excel formulas against Per-Image columns H and K.
    # Per-Image sheet layout (1-indexed columns):
    #   A=Index, B=ImageID, C=FileName, D=W, E=H,
    #   F=TFLite#, G=TFLiteMaxScore, H=TFLiteLatencyMs,
    #   I=PT#,     J=PTMaxScore,     K=PTLatencyMs.
    last_data_row = per_image_count + 1  # header is row 1
    tf_range = f"'Per-Image'!H2:H{last_data_row}"
    pt_range = f"'Per-Image'!K2:K{last_data_row}"

    row += 1
    sheet.cell(row=row, column=1, value="Latency").font = _LABEL_FONT
    sheet.cell(row=row, column=2, value="TFLite (ms)").font = _LABEL_FONT
    sheet.cell(row=row, column=3, value="PyTorch (ms)").font = _LABEL_FONT
    row += 1

    latency_stats: Sequence[tuple[str, str]] = (
        ("Mean",   "AVERAGE"),
        ("Median", "MEDIAN"),
        ("Min",    "MIN"),
        ("Max",    "MAX"),
        ("Stdev",  "STDEV"),
    )
    for label, fn in latency_stats:
        sheet.cell(row=row, column=1, value=label).font = _LABEL_FONT
        sheet.cell(row=row, column=2, value=f"={fn}({tf_range})").number_format = "0.00"
        sheet.cell(row=row, column=3, value=f"={fn}({pt_range})").number_format = "0.00"
        row += 1

    _autosize_columns(sheet)


def _write_per_class_sheet(
    sheet: Worksheet,
    *,
    tflite_per_class: dict[int, float],
    pytorch_per_class: dict[int, float],
) -> None:
    headers = (
        "VOC Index",
        "VOC Name",
        "COCO Category ID",
        "TFLite AP@0.5",
        "PyTorch AP@0.5",
        "Delta (TFLite - PyTorch)",
    )
    for col, h in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=h)
    _style_header_row(sheet, 1, len(headers))

    for row_idx, voc_idx in enumerate(
        sorted(VOC_INDEX_TO_COCO_ID.keys()), start=2
    ):
        coco_id = VOC_INDEX_TO_COCO_ID[voc_idx]
        sheet.cell(row=row_idx, column=1, value=voc_idx)
        sheet.cell(row=row_idx, column=2, value=VOC_INDEX_NAMES[voc_idx])
        sheet.cell(row=row_idx, column=3, value=coco_id)
        sheet.cell(
            row=row_idx, column=4,
            value=tflite_per_class.get(coco_id, 0.0),
        ).number_format = "0.0000"
        sheet.cell(
            row=row_idx, column=5,
            value=pytorch_per_class.get(coco_id, 0.0),
        ).number_format = "0.0000"
        sheet.cell(
            row=row_idx, column=6,
            value=f"=D{row_idx}-E{row_idx}",
        ).number_format = "0.0000;[Red]-0.0000"

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _write_per_image_sheet(
    sheet: Worksheet, per_image_rows: list[dict[str, Any]]
) -> None:
    headers = (
        "Index", "Image ID", "File Name", "Width", "Height",
        "TFLite # Dets", "TFLite Max Score", "TFLite Latency (ms)",
        "PyTorch # Dets", "PyTorch Max Score", "PyTorch Latency (ms)",
    )
    for col, h in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=h)
    _style_header_row(sheet, 1, len(headers))

    for row_idx, r in enumerate(per_image_rows, start=2):
        sheet.cell(row=row_idx, column=1,  value=r["index"])
        sheet.cell(row=row_idx, column=2,  value=r["image_id"])
        sheet.cell(row=row_idx, column=3,  value=r["file_name"])
        sheet.cell(row=row_idx, column=4,  value=r["width"])
        sheet.cell(row=row_idx, column=5,  value=r["height"])
        sheet.cell(row=row_idx, column=6,  value=r["tflite_detections"])
        sheet.cell(row=row_idx, column=7,  value=r["tflite_max_score"]).number_format = "0.0000"
        sheet.cell(row=row_idx, column=8,  value=r["tflite_latency_ms"]).number_format = "0.00"
        sheet.cell(row=row_idx, column=9,  value=r["pytorch_detections"])
        sheet.cell(row=row_idx, column=10, value=r["pytorch_max_score"]).number_format = "0.0000"
        sheet.cell(row=row_idx, column=11, value=r["pytorch_latency_ms"]).number_format = "0.00"

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _write_config_sheet(
    sheet: Worksheet,
    *,
    tflite_model_path: Path,
    pytorch_model_path: Path,
    coco_gt_json: Path,
    coco_images_dir: Path,
    score_threshold: float,
    num_images_evaluated: int,
) -> None:
    sheet["A1"] = "Run Configuration"
    sheet["A1"].font = _TITLE_FONT
    sheet.merge_cells("A1:B1")

    rows: Sequence[tuple[str, Any]] = (
        ("TFLite model",      str(tflite_model_path)),
        ("PyTorch model",     str(pytorch_model_path)),
        ("COCO GT JSON",      str(coco_gt_json)),
        ("COCO images dir",   str(coco_images_dir)),
        ("Score threshold",   score_threshold),
        ("Images evaluated",  num_images_evaluated),
    )
    for row_idx, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_idx, column=1, value=label).font = _LABEL_FONT
        sheet.cell(row=row_idx, column=2, value=value)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90


# ---------------------------------------------------------------------------
# Comparison driver.
# ---------------------------------------------------------------------------

def run_comparison(
    *,
    coco_gt_json: Path,
    coco_images_dir: Path,
    tflite_model_path: Path,
    pytorch_model_path: Path,
    qfgaohao_repo_path: Path,
    output_xlsx: Path,
    num_images: int,
    score_threshold: float,
    debug_images: int = 0,
    debug_dir: Optional[Path] = None,
    debug_display_score_threshold: float = 0.3,
) -> None:
    """Run both evaluators over ``num_images`` COCO val2017 images and write xlsx.

    If ``debug_images > 0``, the first ``debug_images`` samples also get a
    3-panel annotated PNG (GT | TFLite | qfgaohao) written under ``debug_dir``
    (defaults to ``<output_xlsx parent>/debug/``).
    """
    logger.info("Loading COCO GT: %s", coco_gt_json)
    coco_gt = COCO(str(coco_gt_json))

    image_ids = sorted(coco_gt.getImgIds())[:num_images]
    logger.info("Evaluating on %d images", len(image_ids))

    tflite_eval = TFLiteSSDEvaluator(tflite_model_path)
    pytorch_eval = PyTorchSSDEvaluator(pytorch_model_path, qfgaohao_repo_path)

    # Build a category_id -> name lookup once for debug labels.
    category_name_by_id: dict[int, str] = {
        int(cid): coco_gt.cats[cid].get("name", str(cid))
        for cid in coco_gt.getCatIds()
    }

    if debug_images > 0:
        debug_dir = debug_dir or (output_xlsx.parent / "debug")
        debug_dir.mkdir(parents=True, exist_ok=True)
        logger.info(
            "Will render %d debug images to %s (display threshold %.2f)",
            debug_images, debug_dir, debug_display_score_threshold,
        )

    tflite_predictions: list[dict[str, Any]] = []
    pytorch_predictions: list[dict[str, Any]] = []
    per_image_rows: list[dict[str, Any]] = []

    progress_every = max(1, len(image_ids) // 20)

    for idx, image_id in enumerate(image_ids):
        info = coco_gt.loadImgs([image_id])[0]
        image_path = coco_images_dir / info["file_name"]
        if not image_path.exists():
            logger.warning("Missing image: %s; skipping", image_path)
            continue
        try:
            pil_image = Image.open(image_path).convert("RGB")
        except (OSError, ValueError) as exc:
            logger.warning("Failed to open %s: %s; skipping", image_path, exc)
            continue

        width, height = pil_image.size

        tf_dets, tf_ms = tflite_eval.detections_for_image(
            pil_image, image_id, width, height, score_threshold
        )
        pt_dets, pt_ms = pytorch_eval.detections_for_image(
            pil_image, image_id, width, height, score_threshold
        )

        tflite_predictions.extend(tf_dets)
        pytorch_predictions.extend(pt_dets)

        per_image_rows.append({
            "index":               idx,
            "image_id":            image_id,
            "file_name":           info["file_name"],
            "width":               width,
            "height":              height,
            "tflite_detections":   len(tf_dets),
            "tflite_max_score":    max((d["score"] for d in tf_dets), default=0.0),
            "tflite_latency_ms":   tf_ms,
            "pytorch_detections":  len(pt_dets),
            "pytorch_max_score":   max((d["score"] for d in pt_dets), default=0.0),
            "pytorch_latency_ms":  pt_ms,
        })

        if debug_images > 0 and idx < debug_images and debug_dir is not None:
            gt_anns = coco_gt.loadAnns(coco_gt.getAnnIds(imgIds=[image_id]))
            debug_path = debug_dir / f"{idx:04d}_img{image_id:012d}.png"
            try:
                render_debug_image(
                    pil_image,
                    gt_anns=gt_anns,
                    tflite_dets=tf_dets,
                    pytorch_dets=pt_dets,
                    category_name_by_id=category_name_by_id,
                    save_path=debug_path,
                    display_score_threshold=debug_display_score_threshold,
                )
            except Exception as exc:
                logger.warning("Failed to render debug image %s: %s", debug_path, exc)

        if (idx + 1) % progress_every == 0:
            logger.info(
                "  [%d/%d] tflite=%.1fms pytorch=%.1fms",
                idx + 1, len(image_ids), tf_ms, pt_ms,
            )

    # COCOeval passes.
    logger.info("Running COCOeval: TFLite full COCO 80...")
    tflite_full, _ = run_coco_eval(
        coco_gt, tflite_predictions, image_ids=image_ids
    )

    logger.info("Running COCOeval: TFLite VOC-equivalent subset...")
    tflite_voc, tflite_per_class = run_coco_eval(
        coco_gt, tflite_predictions,
        category_ids=VOC_EQUIVALENT_COCO_IDS,
        image_ids=image_ids,
    )

    logger.info("Running COCOeval: PyTorch (qfgaohao) VOC-equivalent subset...")
    pytorch_voc, pytorch_per_class = run_coco_eval(
        coco_gt, pytorch_predictions,
        category_ids=VOC_EQUIVALENT_COCO_IDS,
        image_ids=image_ids,
    )

    logger.info("Writing xlsx: %s", output_xlsx)
    write_xlsx(
        output_xlsx,
        num_images_evaluated=len(per_image_rows),
        tflite_full_metrics=tflite_full,
        tflite_voc_metrics=tflite_voc,
        pytorch_voc_metrics=pytorch_voc,
        tflite_per_class=tflite_per_class,
        pytorch_per_class=pytorch_per_class,
        per_image_rows=per_image_rows,
        tflite_model_path=tflite_model_path,
        pytorch_model_path=pytorch_model_path,
        coco_gt_json=coco_gt_json,
        coco_images_dir=coco_images_dir,
        score_threshold=score_threshold,
    )

    # Console summary.
    bar = "=" * 70
    logger.info(bar)
    logger.info("FINAL SUMMARY (VOC-equivalent 20-class subset)")
    logger.info(bar)
    logger.info(
        "TFLite  mAP@0.5: %.4f | mAP@[0.5:0.95]: %.4f",
        tflite_voc["mAP_0.5"], tflite_voc["mAP_0.5_0.95"],
    )
    logger.info(
        "PyTorch mAP@0.5: %.4f | mAP@[0.5:0.95]: %.4f",
        pytorch_voc["mAP_0.5"], pytorch_voc["mAP_0.5_0.95"],
    )
    if per_image_rows:
        tflite_mean = float(np.mean([r["tflite_latency_ms"] for r in per_image_rows]))
        pytorch_mean = float(np.mean([r["pytorch_latency_ms"] for r in per_image_rows]))
        logger.info("TFLite  avg latency: %.2f ms", tflite_mean)
        logger.info("PyTorch avg latency: %.2f ms", pytorch_mean)


# ---------------------------------------------------------------------------
# CLI.
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    """Parse CLI args, defaulting paths to the MetaExecuTorch project layout."""
    parser = argparse.ArgumentParser(
        description=(
            "Compare TFLite vs PyTorch (qfgaohao) MobileNetV2-SSD on COCO val2017."
        )
    )

    # __file__ is expected to live under
    # MetaExecuTorch/executorch-toolkit/evaluation/tflite_comparison/.
    # parents[3] is MetaExecuTorch/. Falls back to cwd if outside that layout.
    here = Path(__file__).resolve()
    if len(here.parents) >= 4:
        project_root = here.parents[3]
    else:
        project_root = Path.cwd()

    parser.add_argument(
        "--tflite-model", type=Path,
        default=(
            project_root / "model_sources" / "MobileNetV2_TFLite"
            / "tf2_ssd_mobilenet_v2_coco17_ptq.tflite"
        ),
        help="Path to the .tflite model file.",
    )
    parser.add_argument(
        "--pytorch-model", type=Path,
        default=(
            project_root / "model_sources" / "MobileNetV2" / "weights"
            / "mobile_net_v2_ssd.pth"
        ),
        help="Path to the qfgaohao MobileNetV2-SSD-Lite .pth checkpoint.",
    )
    parser.add_argument(
        "--qfgaohao-repo", type=Path,
        default=(
            project_root / "model_sources" / "MobileNetV2" / "src"
            / "pytorch" / "pytorch-ssd"
        ),
        help="Path to the qfgaohao pytorch-ssd source repo.",
    )
    parser.add_argument(
        "--coco-images", type=Path,
        default=project_root / "dataset" / "coco_val2017" / "val2017",
        help="Directory containing COCO val2017 .jpg images.",
    )
    parser.add_argument(
        "--coco-annotations", type=Path,
        default=(
            project_root / "dataset" / "coco_val2017" / "annotations"
            / "instances_val2017.json"
        ),
        help="Path to the COCO val2017 instances JSON.",
    )
    parser.add_argument(
        "--output-xlsx", type=Path,
        default=(
            project_root / "output" / "comparisons"
            / "tflite_vs_pytorch_coco.xlsx"
        ),
        help="Where to write the comparison xlsx.",
    )
    parser.add_argument(
        "--num-images", type=int, default=1000,
        help="Number of COCO val images to evaluate (sorted by image_id).",
    )
    parser.add_argument(
        "--score-threshold", type=float, default=0.01,
        help="Drop detections with score below this threshold.",
    )
    parser.add_argument(
        "--debug-images", type=int, default=20,
        help=(
            "Render this many annotated debug PNGs (GT | TFLite | qfgaohao) "
            "for sanity-checking boxes. Set 0 to disable."
        ),
    )
    parser.add_argument(
        "--debug-score-threshold", type=float, default=0.3,
        help=(
            "Display-only score threshold for detection boxes in debug PNGs "
            "(separate from --score-threshold which drives mAP)."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    # Validate all input paths up front.
    required_paths: Sequence[tuple[str, Path]] = (
        ("TFLite model",     args.tflite_model),
        ("PyTorch model",    args.pytorch_model),
        ("qfgaohao repo",    args.qfgaohao_repo),
        ("COCO images dir",  args.coco_images),
        ("COCO annotations", args.coco_annotations),
    )
    missing = [(label, p) for label, p in required_paths if not p.exists()]
    if missing:
        for label, path in missing:
            logger.error("%s not found: %s", label, path)
        return 1

    run_comparison(
        coco_gt_json=args.coco_annotations,
        coco_images_dir=args.coco_images,
        tflite_model_path=args.tflite_model,
        pytorch_model_path=args.pytorch_model,
        qfgaohao_repo_path=args.qfgaohao_repo,
        output_xlsx=args.output_xlsx,
        num_images=args.num_images,
        score_threshold=args.score_threshold,
        debug_images=args.debug_images,
        debug_display_score_threshold=args.debug_score_threshold,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
